from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


FIELDS = [
    "student_id",
    "name",
    "course",
    "year_level",
    "age",
    "gender",
    "contact_number",
    "address",
    "emergency_name",
    "emergency_relationship",
    "emergency_contact_number",
    "status",
]
HEADERS = [
    "Student ID",
    "Name",
    "Course",
    "Year",
    "Age",
    "Gender",
    "Contact Number",
    "Address",
    "Emergency Contact Name",
    "Emergency Contact Relationship",
    "Emergency Contact Number",
    "Status",
]
COLUMN_WIDTHS = {
    "A": 18,
    "B": 28,
    "C": 58,
    "D": 14,
    "E": 8,
    "F": 14,
    "G": 18,
    "H": 38,
    "I": 28,
    "J": 34,
    "K": 28,
    "L": 14,
}
LEGACY_HEADERS = ["Student ID", "Name", "Course", "Year", "Contact Number", "Status"]
LEGACY_FIELDS = ["student_id", "name", "course", "year_level", "contact_number", "status"]
AGE_HEADERS = ["Student ID", "Name", "Course", "Year", "Age", "Contact Number", "Status"]
AGE_FIELDS = ["student_id", "name", "course", "year_level", "age", "contact_number", "status"]
ADDRESS_HEADERS = ["Student ID", "Name", "Course", "Year", "Age", "Contact Number", "Address", "Status"]
ADDRESS_FIELDS = ["student_id", "name", "course", "year_level", "age", "contact_number", "address", "status"]
GENDER_HEADERS = ["Student ID", "Name", "Course", "Year", "Age", "Gender", "Contact Number", "Address", "Emergency Contact Name", "Emergency Contact Relationship", "Emergency Contact Number", "Status"]
GENDER_FIELDS = FIELDS


class StudentModel:
    def __init__(self, file_name="students.xlsx"):
        self.file_path = Path(file_name)
        self.legacy_json_path = Path("students.json")
        self.ensure_workbook()
        self.students = self.load_students()

    def ensure_workbook(self):
        if self.file_path.exists():
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Students"
        sheet.append(HEADERS)
        format_student_sheet(sheet)

        legacy_students = self.load_legacy_json()
        for student in legacy_students:
            sheet.append([student.get(field, "") for field in FIELDS])

        workbook.save(self.file_path)

    def load_legacy_json(self):
        if not self.legacy_json_path.exists():
            return []

        try:
            import json

            with self.legacy_json_path.open("r") as file:
                data = json.load(file)
        except (OSError, json.JSONDecodeError):
            return []

        if not isinstance(data, list):
            return []

        return [student for student in data if isinstance(student, dict)]

    def load_students(self):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active
        students = []
        headers = [cell.value for cell in sheet[1]]
        if headers == LEGACY_HEADERS:
            fields = LEGACY_FIELDS
        elif headers == AGE_HEADERS:
            fields = AGE_FIELDS
        elif headers == ADDRESS_HEADERS:
            fields = ADDRESS_FIELDS
        elif headers == GENDER_HEADERS:
            fields = GENDER_FIELDS
        else:
            fields = FIELDS

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            student = {}
            for field, value in zip(fields, row):
                student[field] = "" if value is None else str(value)

            student.setdefault("age", "")
            student.setdefault("gender", "")
            student.setdefault("address", "")
            student.setdefault("emergency_name", "")
            student.setdefault("emergency_relationship", "")
            student.setdefault("emergency_contact_number", "")
            student.setdefault("status", "Active")
            students.append(student)

        return students

    def save_students(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Students"
        sheet.append(HEADERS)

        for student in self.students:
            sheet.append([student.get(field, "") for field in FIELDS])

        format_student_sheet(sheet)

        try:
            workbook.save(self.file_path)
        except PermissionError:
            return False

        return True

    def get_all_students(self):
        return self.students

    def find_student(self, student_id):
        for student in self.students:
            if student["student_id"] == student_id:
                return student
        return None

    def add_student(self, student):
        if self.find_student(student["student_id"]):
            return False

        self.students.append(student)
        if not self.save_students():
            self.students.remove(student)
            return "locked"

        return True

    def search_students(self, query):
        query = query.strip().lower()
        if not query:
            return self.students

        return [
            student
            for student in self.students
            if query in student["student_id"].lower()
            or query in student["name"].lower()
            or query in student["course"].lower()
            or query in student["year_level"].lower()
            or query in student["contact_number"].lower()
            or query in student.get("status", "").lower()
        ]

    def update_student(self, student_id, updates):
        student = self.find_student(student_id)
        if not student:
            return False

        original = student.copy()
        for field, value in updates.items():
            if value:
                student[field] = value

        if not self.save_students():
            student.clear()
            student.update(original)
            return "locked"

        return True

    def delete_student(self, student_id):
        student = self.find_student(student_id)
        if not student:
            return False

        self.students.remove(student)
        if not self.save_students():
            self.students.append(student)
            return "locked"

        return True


def format_student_sheet(sheet):
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_fill = PatternFill("solid", fgColor="DFF7FB")
    header_font = Font(bold=True, color="12313A")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for column_letter, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[column_letter].width = width

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_alignment

    sheet.row_dimensions[1].height = 32
